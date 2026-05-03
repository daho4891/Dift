from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from dift.reports.models import DiffReport


def render_excel(report: DiffReport, output: str | None = None) -> Path:
    """
    Render and write an Excel report.

    Excel reports should always be written to a file.
    """

    output_path = Path(output or "dift_report.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    _add_summary_sheet(ws, report)

    wb.save(output_path)
    return output_path


def _add_summary_sheet(ws, report: DiffReport) -> None:
    rows = [
        ["Metric", "Value"],
        ["old_rows", report.summary.old_rows],
        ["new_rows", report.summary.new_rows],
        ["row_delta", report.summary.row_delta],
        ["risk_level", report.summary.risk_level],
    ]

    for row in rows:
        ws.append(row)

    _style_header(ws)
    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font


def _auto_size_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)